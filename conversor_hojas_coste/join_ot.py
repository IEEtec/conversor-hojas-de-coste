import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def juntar_ot_en_workbook(
    ws_materiales: Worksheet,
    ws_mano_obra: Worksheet,
    fichero: str,
    errores: bool,
    verboso: bool,
) -> None:

    # Abre el fichero OT y añade sus datos al workbook dado

    wb_origen = load_workbook(fichero)

    # Procesa la hoja de materiales

    ws_materiales_origen = wb_origen["Materiales"]
    ws_mano_obra_origen = wb_origen["Mano de Obra"]

    for row in ws_materiales_origen.iter_rows(
        min_row=2, min_col=1, values_only=True
    ):
        # Las primeras 4 columnas se pueden copiar tal cual, la quinta columna (E) (Fecha), debe convertirse a datetime y posteriormente a date para eliminar la hora. Las columnas F a J son numéricas y se pueden copiar tal cual.

        # Añade una nueva fila vacia a la hoja de materiales
        ws_materiales.append(row)

    for cell in ws_materiales["E"]:
        cell.number_format = "dd/mm/yyyy"

    number_format = "#,##0.00#"

    # Iterate over all cells in columns F to J and set the number format

    for cell in (
        ws_materiales["F"]
        + ws_materiales["G"]
        + ws_materiales["H"]
        + ws_materiales["I"]
        + ws_materiales["J"]
    ):
        cell.number_format = number_format

    for row in ws_mano_obra_origen.iter_rows(
        min_row=2, min_col=1, values_only=True
    ):
        # Procesa cada fila de mano de obra
        ws_mano_obra.append(row)

    for cell in ws_materiales["E"]:
        cell.number_format = "dd/mm/yyyy"

    for cell in (
        ws_mano_obra["H"]
        + ws_mano_obra["I"]
        + ws_mano_obra["J"]
        + ws_mano_obra["K"]
        + ws_mano_obra["L"]
        + ws_mano_obra["M"]
        + ws_mano_obra["N"]
    ):
        cell.number_format = number_format
