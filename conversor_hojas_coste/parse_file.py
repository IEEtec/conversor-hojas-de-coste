import re

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from material_line import MaterialLine
from job_line import JobLine


# A partir de dos archivos, uno con un nombre como "nnnnnn.txt" y otro con un nombre como "nnnnnncoste.txt", donde n es un número natural
# Este script analizará los archivos y creará nuevos archivos llamados "nnnnnn_materiales.csv" y "nnnnnn_trabajo.csv"
# La codificación del archivo es windows-1252

# El archivo tiene este formato:

#    Hoja de Coste Nº 230408

#  INSTALACIONES ELEC. EGARA, SL                                                                                      jueves, 20 de marzo de 2025
#                   - AMPLIACIÓN TUNEL --> 2.000,00€

# _________________________________________________M A T E R I A L_________________________________________________
# Referencia       Descripción                              Fecha         Cantidad          Precio         Importe
# _________________________________________________________________________________________________________________
#                  PASAMUROS ICOTEK KEL-DPU 24/20 GRIS      09/10/2024       10,00           44,03          440,30
# CAB1,5LILA       CABLE LILA 1,5 CPR LH CCA                18/06/2024       40,00            0,20            8,00
# PHO3113771       TAPA FINAL D-PIT 1,5/S-3L GY             25/11/2024      -27,00            0,53          -14,31
# PHO3113771       TAPA FINAL D-PIT 1,5/S-3L GY             08/05/2024       50,00            0,53           26,50

#                                                                                     Pág.     1
#                   - ENTREGA DE LOS ESQUEMAS ELÉCTRICOS --> SEMANA 23 4% --> 1.526,42€
#                   - ENTREGA DE LA PROGRAMACIÓN PLC & HMI --> SEMANA 24 15% --> 5.400,00€
#                   - AMPLIACIÓN TUNEL --> 2.000,00€
# PHO3208100       BORNA CONEIXION PIT-1,5/S                03/05/2024       50,00            0,61           30,50
# PHO3213056       PUENTE ENCHUFABLE FBS 10-3,5             08/05/2024       50,00            3,03          151,50
# .....

# SIE6ES72141HG400 SIMATIC S7-1200, CPU 1214C               14/03/2025       -1,00         327,479         -327,48
# WEI1240840000    SWITCH DE RED NO GESTIONADO IE-SW        31/01/2025        3,00          63,420          190,26
#                                                                      ___________                 _______________
#                                                                         1.677,10                       14.845,22

# _____________________________________________M A N O   D E   O B R A_____________________________________________
# Operaci�n                     Fecha      Operario                        Cant.  Precio   Dietas Despl.    Importe
# _________________________________________________________________________________________________________________
#    1 COLABORACION CON ALVARO  20/11/2024   96 RIVERA ARISTIZABAL, JORG    1,00  25,000     0,00   0,00      25,00
#    1 General                  19/12/2024    9 L�PEZ NIETO, ALVARO         8,00  25,000     0,00   0,00     200,00

# Vamos a cargar el archivo completo en memoria y luego lo vamos a parsear para extraer la información que necesitamos


# Primero, podemos eliminar las primeras líneas hasta encontrar una línea como esta:
# _________________________________________________M A T E R I A L_________________________________________________

# Descartar también las siguientes dos líneas
# Después podemos dividir el archivo en líneas y extraer la información que necesitamos

# Vamos a crear una clase para almacenar la información del archivo, MaterialLine, y crear una lista
# Parsear una línea del archivo y almacenar la información en un objeto MaterialLine y añadirlo a la lista
# Tras procesar la línea de material, hay que comprobar el primer carácter de la línea; si es un espacio, el campo Referencia está vacío
# Si es otro salto de línea (CRLF), es el final de la página, por lo que hay que omitir líneas hasta que, comprobando el primer carácter, encontremos una línea con un carácter distinto de espacio
# Si el carácter es un guion bajo, significa que hemos llegado al final de la sección de materiales, así que hay que omitir la siguiente línea y empezar a leer la sección de mano de obra

# Aquí vamos a intentar parsear la primera parte del archivo, la sección de materiales


def ParseFile(
    filename_cost: str,
    filename_sell: str,
    numero_ot: str,
    errors: bool,
    verbose: bool,
) -> None:

    line_number = 0

    lines_cost: list[str] = []
    lines_sell: list[str] = []

    if filename_cost != "":
        with open(filename_cost, "r", encoding="UTF-16LE") as file:
            lines_cost = file.readlines()

    if filename_sell != "":
        with open(filename_sell, "r", encoding="UTF-16LE") as file:
            lines_sell = file.readlines()

    lines_guide: list[str] = []

    if len(lines_cost) > 0:
        lines_guide = lines_cost

    elif len(lines_sell) > 0:
        lines_guide = lines_sell

    else:
        raise ValueError("No se han proporcionado archivos para parsear.")

    use_cost = len(lines_cost) > 0
    use_sell = len(lines_sell) > 0

    if verbose:
        print(f"Usando coste: {use_cost}, usando venta: {use_sell}")
        print(f"Número de líneas en coste: {len(lines_cost)}")
        print(f"Número de líneas en venta: {len(lines_sell)}")
        print(f"Número de líneas en guía: {len(lines_guide)}")

    # En la fila 6 se puede ver el rango de fechas de la hoja de coste
    # Desde Fecha  20/03/2024      Hasta Fecha  20/03/2025

    fecha_desde = None
    fecha_hasta = None

    match_fecha = re.search(
        r"Desde Fecha\s+(\d{2}/\d{2}/\d{4})\s+Hasta Fecha\s+(\d{2}/\d{2}/\d{4})",
        lines_guide[5],
    )

    if match_fecha:
        fecha_desde = match_fecha.group(1)
        fecha_hasta = match_fecha.group(2)
        print(f"Fecha desde: {fecha_desde}, Fecha hasta: {fecha_hasta}")
    else:
        print("No se encontró el rango de fechas.")
        print(lines_guide[5])
        raise ValueError(
            "No se pudo encontrar el rango de fechas en la línea 5."
        )

    fecha_desde = fecha_desde.replace("/", "")
    fecha_hasta = fecha_hasta.replace("/", "")

    # Remove the first lines until the first character of the line is not other than a "_". It must be the first character.

    while lines_guide[line_number][0] != "_":
        line_number += 1

    # Comprueba que la linea actual es el header de la seccion de materiales
    # _________________________________________________M A T E R I A L_________________________________________________

    if (
        re.match(r"^_+M A T E R I A L_+$", lines_guide[line_number].strip())
        == None
    ):
        raise ValueError(
            f"La línea {line_number} no es el encabezado de la sección de materiales: {lines_guide[line_number]}"
        )

    # Nos saltamos esta y las dos siguientes líneas

    line_number += 2

    # La linea actual es la primera línea de material

    # Parsear las líneas de material hasta encontrar el final de la primera pagina o el final de la sección de materiales

    # Este es un ejemplo del final de una pagina:

    # RIT3110000       TERMOSTATO SK PARA INTERRUPTOR ARMARIO   25/11/2024        1,00          29,893           29,89
    #
    #                                                                                    P�g.     2
    #   Hoja de Coste N� 230582
    #
    # IN
    #
    # STALACIONES ELEC. EGARA, SL                                                                                      jueves, 20 de marzo de 2025
    #
    #      Cliente     230 AD-TECH IBERICA, SL                                  Responsable
    #      Desde Fecha  19/09/2024      Hasta Fecha  20/03/2025                 Valoraci�n Coste
    #      Descripci�n LIFTED NOVORDISK
    #                  Pto. n� 4000290
    #                  Cuadros el�ctricos -->          36.573,00�
    #                  Esquemas el�ctricos -->         3.200,00�
    #                  Programaci�n PLC + HMI --> 6.400,00�
    # RIT3240200       FILTRO SALIDA PARA 3240 3241 RAL7035     02/12/2024        1,00          32,215           32,22

    # La linea esta vacia completamente. Al detectar una linea con solo un salto de linea vamos a empezar a saltar lineas hasta encontrar una linea que el primer caracter no sea tambien un espacio o un salto de linea

    # El final de la seccion de materiales es una linea como esta:

    # WEI1240840000    SWITCH DE RED NO GESTIONADO IE-SW        31/01/2025        3,00          63,420          190,26
    #                                                                     ___________                 _______________
    #                                                                        1.677,10                       14.845,22

    # Este caso lo veremos porque es una linea que tiene un espacio en la primera posicion.

    # Create a list to store the material lines

    material_lines: list[MaterialLine] = []

    # En resumen. Iterar sobre las líneas

    # Si el primer carácter no es un espacio ni un salto de línea (CRLF), es una línea de material: parsearla y almacenar la información en un objeto MaterialLine
    # Si es un salto de línea (CRLF), es el final de la página; hay que saltar líneas hasta encontrar una cuyo primer carácter no sea espacio o salto de línea
    # Si es un espacio, puede ser que el campo Referencia esté vacío o que sea el final de la lista
    # Si es un guion bajo, es el final de la sección de materiales; terminar el bucle
    # Si es una letra, continuar parseando el archivo

    while line_number < len(lines_guide):

        line_number += 1

        if use_cost:
            line_cost = lines_cost[line_number]
        else:
            line_cost = ""

        if use_sell:
            line_sell = lines_sell[line_number]
        else:
            line_sell = ""

        line_guide = lines_guide[line_number]

        if verbose:
            if use_cost:
                print(
                    f"Linea material coste {line_number}: {line_cost}", end=""
                )
            if use_sell:
                print(
                    f"Linea material venta {line_number}: {line_sell}", end=""
                )
            print(f"Linea material guia {line_number}: {line_guide}", end="")

        # Si la linea empieza por un salto de linea, ignorarla
        if line_guide[0] == "\n":
            continue

        # Si la linea es exactamente:
        #                                                                     ___________                 _______________
        # Hemos llegado al final de la seccion de materiales

        if (
            line_guide
            == "                                                                     ___________                 _______________\n"
        ):
            break

        new_material_line: MaterialLine | None = None

        try:
            new_material_line = MaterialLine.parse(
                line_guide,
                line_cost,
                line_sell,
                line_number,
                use_cost,
                use_sell,
            )
        except MaterialLine.NotValidError as e:
            if errors:
                print(e)
            pass
        except MaterialLine.IncongruentError as e:
            print(e)
        if new_material_line is not None:
            if verbose:
                print(f"New material line: {new_material_line}")
            material_lines.append(new_material_line)

    # Ahora, saltar las líneas hasta encontrar el inicio de la sección de mano de obra
    while line_number < len(lines_guide):

        line_number += 1

        line = lines_guide[line_number]

        if line[0] == "_":
            # Comprobar si la siguiente linea es el encabezado de la seccion de mano de obra
            if re.match(
                r"^_+M A N O   D E   O B R A_+$",
                lines_guide[line_number].strip(),
            ):
                # Hemos llegado al inicio de la seccion de mano de obra
                break

    line_number += 2

    #   1 ARMADO DE CUADRO DE DINA 20/08/2024  109 GARCIA RODRIGUEZ, KEVIN     3,00   30,00     0,00   0,00      90,00

    # Now, until the end, parse the lines that have a "1" in the fourth position

    # Create a list to store the job lines

    job_lines: list[JobLine] = []

    # Iterate over the lines

    while line_number < len(lines_guide):

        line_number += 1

        if use_cost:
            line_cost = lines_cost[line_number]
        else:
            line_cost = ""

        if use_sell:
            line_sell = lines_sell[line_number]
        else:
            line_sell = ""

        line_guide = lines_guide[line_number]

        if verbose:
            if use_cost:
                print(f"Linea job coste {line_number}: {line_cost}", end="")
            if use_sell:
                print(f"Linea job venta {line_number}: {line_sell}", end="")
            print(f"Linea job guia {line_number}: {line_guide}", end="")

        # Si la linea es la siguiente:                                                                     __________         ________ ______ __________

        # Si la linea es exactamente esta, hemos llegado al final de la seccion de mano de obra

        if (
            line_guide
            == "                                                                    __________         ________ ______ __________\n"
        ):
            break

        new_job_line: JobLine | None = None

        try:
            new_job_line = JobLine.parse(
                line_guide,
                line_cost,
                line_sell,
                line_number,
                use_cost,
                use_sell,
            )
        except JobLine.NotValidError as e:
            if errors:
                print(e)
        except JobLine.IncongruentError as e:
            print(e)

        if new_job_line is not None:
            if verbose:
                print(f"New job line: {new_job_line}")
            job_lines.append(new_job_line)

    print(f"Se han parseado {len(material_lines)} líneas de material.")
    print(f"Se han parseado {len(job_lines)} líneas de mano de obra.")

    # Now we are going to write the information to a csv file

    filename_material = numero_ot + "_materiales.xlsx"
    filename_job = numero_ot + "_manodeobra.xlsx"
    filename = f"{numero_ot}_de_{fecha_desde}_a_{fecha_hasta}.xlsx"

    # Write the material lines

    wb = Workbook()
    ws: Worksheet = wb.active
    if ws is None:
        raise ValueError("No se pudo crear la hoja de cálculo.")

    ws.title = "Materiales"

    ws.append(
        [
            "Linea fichero original",
            "Numero OT",
            "Referencia",
            "Descripcion",
            "Fecha",
            "Cantidad",
            "Precio Coste",
            "Importe Coste",
            "Precio Venta",
            "Importe Venta",
        ]
    )

    for material_line in material_lines:

        ws.append(
            [
                material_line.line_number + 1,
                numero_ot,
                material_line.Referencia,
                material_line.Descripcion,
                material_line.Fecha,
                material_line.Cantidad,
                material_line.PrecioUnitarioCoste,
                material_line.ImporteTotalCoste,
                material_line.PrecioUnitarioVenta,
                material_line.ImporteTotalVenta,
            ]
        )

    ws.column_dimensions["A"].width = 19
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 14

    ws.row_dimensions[1].height = 36
    ws.row_dimensions[1].alignment = Alignment(wrap_text=True)

    number_format = "#,##0.00#"

    # Iterate over all cells in columns F to J and set the number format

    for cell in ws["F"] + ws["G"] + ws["H"] + ws["I"] + ws["J"]:
        cell.number_format = number_format

    for cell in ws["E"]:
        cell.number_format = "dd/mm/yyyy"

    table = Table(
        displayName="materiales", ref="A1:J" + str(len(material_lines) + 1)
    )

    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )

    table.tableStyleInfo = style
    ws.add_table(table)

    # Write the job lines

    ws = wb.create_sheet(title="Mano de Obra")

    ws.append(
        [
            "Linea fichero original",
            "Numero OT",
            "Operacion Id",
            "Operacion",
            "Fecha",
            "Operario Id",
            "Operario Nombre",
            "Cantidad",
            "Precio Coste",
            "Importe Coste",
            "Dietas Coste",
            "Desplazamiento Coste",
            "Precio Venta",
            "Importe Venta",
            "Dietas Venta",
            "Desplazamiento Venta",
        ]
    )

    for job_line in job_lines:
        ws.append(
            [
                job_line.line_number + 1,
                numero_ot,
                job_line.OperacionId,
                job_line.Operacion,
                job_line.Fecha,
                job_line.OperarioId,
                job_line.OperarioNombre,
                job_line.Cantidad,
                job_line.PrecioUnitarioCoste,
                job_line.ImporteTotalCoste,
                job_line.DietasCoste,
                job_line.DesplazamientoCoste,
                job_line.PrecioUnitarioVenta,
                job_line.ImporteTotalVenta,
                job_line.DietasVenta,
                job_line.DesplazamientoVenta,
            ]
        )

    ws.column_dimensions["A"].width = 19
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 30
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 14
    ws.column_dimensions["K"].width = 14
    ws.column_dimensions["L"].width = 14
    ws.column_dimensions["M"].width = 14
    ws.column_dimensions["N"].width = 14
    ws.column_dimensions["O"].width = 14
    ws.column_dimensions["P"].width = 14

    ws.row_dimensions[1].height = 36
    ws.row_dimensions[1].alignment = Alignment(wrap_text=True)

    # Iterate over all cells in columns H to N and set the number format

    for cell in (
        ws["H"]
        + ws["I"]
        + ws["J"]
        + ws["K"]
        + ws["L"]
        + ws["M"]
        + ws["N"]
        + ws["O"]
        + ws["P"]
    ):
        cell.number_format = number_format

    for cell in ws["E"]:
        cell.number_format = "dd/mm/yyyy"

    table = Table(
        displayName="mano_obra",
        ref="A1:" + get_column_letter(ws.max_column) + str(ws.max_row),
    )

    ws.add_table(table)

    wb.save(filename)
